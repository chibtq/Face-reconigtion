import datetime

import cv2
import mysql.connector
import os
import openpyxl #thu vien cho phep lam viec voi tep excel
#ket noi voi csdl
mydb = mysql.connector.connect(host='localhost', port = '3306', user='root', password='Chihaha@1983', database='ai')
mycursor = mydb.cursor()
#1tai mo hihnh nhan dien khuon mat da duoc huan luyen tu tap trainingdata.yml
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create() #tạo mô hình nhận diện khuôn mặt sử dụng thuật toán LBP được tích hợp sẵn trong thư viện opencv
recognizer.read('recognizer/trainingData.yml') #mô hình sẽ tự động jọc và tìm rac ác đặc trưng quan trọng từ dữ liệu huấn luyện
subjectName = input("Enter Subject Name: ")

def getProfile(Id):
    sql = "SELECT * FROM people WHERE ID= %s"
    mycursor.execute(sql, (Id,))
    result = mycursor.fetchall()
    profile = None
    for row in result:
        profile = row
    return profile


def exportAttendanceToExcel(profile):
    today = datetime.date.today()
    date_string = today.strftime("%d/%m/%Y").replace("/", "-")
    filename = subjectName + ".xlsx"
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='MSV')
        sheet.cell(row=1, column=2, value='Name')
        workbook.save(filename)
    profile = profile[-2:]
    print(profile)

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    col = sheet.max_column
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        if row[col - 1] != date_string:
            sheet.cell(row=1, column=col + 1, value=date_string)

    # neu sv da co trong file thi them dau x vao cot ngay hom do
    col = sheet.max_column
    found = False
    count = 1
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] == profile[0]:
            found = True
            sheet.cell(row=count, column=col, value='x')
            break
        count += 1
    if not found:
        last_row = sheet.max_row
        sheet.cell(row=last_row + 1, column=1, value=profile[0])
        sheet.cell(row=last_row + 1, column=2, value=profile[1])
        sheet.cell(row=last_row + 1, column=col, value='x')

        # Save the workbook to a file
    workbook.save(filename)

#main
cap = cv2.VideoCapture(0)
font_face = cv2.FONT_HERSHEY_SIMPLEX
font_scale = 1
font_color = (255, 255, 255)
while True:
    ret, img = cap.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        Id, conf = recognizer.predict(gray[y:y + h, x:x + w])
        if conf < 35:
            profile = getProfile(Id)
            exportAttendanceToExcel(profile)
            if profile != None:
                cv2.putText(img, "Name: " + str(profile[2]), (x + 10, y + h + 30), font_face, 1, (0, 255, 0), 2)
                cv2.putText(img, "MSV: " + str(profile[1]), (x + 10, y + h + 60), font_face, 1, (0, 255, 0), 2)
        else:
            cv2.putText(img, "Unknown", (x + 10, y + h + 30), font_face, 1, (0, 255, 0), 2)

    cv2.imshow('img', img)
    if cv2.waitKey(1) & 0xff == ord('q'):
        break
